from PyQt5.QtWidgets import (
    QMainWindow,
    QApplication,
    QLabel,
    QPushButton,
    QFrame,
    QLineEdit,
    QLCDNumber,
    QTextEdit,
    QProgressBar,
    QSpinBox,
    QMenuBar,
    QStatusBar,
    QFileDialog,
    QDialog,
    QVBoxLayout,
    QScrollArea,
    QWidget,
)
from PyQt5.QtGui import QFont
from datetime import datetime, timedelta
import openpyxl as opxl
import random
from PyQt5 import uic
import sys
import time
import logging

logging.basicConfig(
    filename="Logs.txt",
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


class InstructionDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Инструкция")
        self.setGeometry(0, 0, 500, 600)

        layout = QVBoxLayout()

        instruction_text = (
            "Инструкция по использованию программы Chronos\n"
            "\n1. Вам нужно подготовить табели!\n"
            "   1.1 Создайте папку, где будут лежать табели (Например, назовите её Tabels_2077).\n"
            "   1.2 Далее, табели в этой папке переименуйте по образцу "
            "(Tab1.xlsx, Tab2.xlsx и тд), где цифра - это номер месяца.\n"
            "   1.3 Помните. Формат файлов везде должен быть .xlsx !!!\n"
            "\n2. Подготовьте таблицу маршрутного файла!\n"
            "   2.1 Строго придерживайтесь вида образца, который лежит в папке программы!\n"
            "   2.2 Лист, откуда программе брать данные - должен быть первым.\n"
            "   2.3 Не должно быть пустых ячеек напротив операции, иначе программа прекратит действие. "
            "Для решения данной проблемы просто поставьте 0. "
            "Например, В строке операции ОТК в ячейке времени - стоит пустая ячейка. Решение - поставьте 0.\n"
            "\n3. При возможности вносите актуальные данные в файл производственного календаря "
            "(Сейчас в нём есть года 2021 - 2024)."
            "\n     В нём 1 - это нерабочий либо несуществующий день. "
            "Когда вы ставите в ячейку файла календаря цифру 1, то она автоматически закрашивается красным для "
            "наглядности.\n"
            "\n4. ПРИСТУПАЕМ К СОЗДАНИЮ КАРТОЧЕК.\n"
            "   4.1 Шаблон - это файл куда будут вноситься данные карточки. (Внесени происходит по ячейкам)\n"
            "   4.2 Файл операций - основной файл с данными, где есть цеха, шифры, время и тд.\n"
            "   4.3 Календарь - производственный клендарь. Нужен Для автовыборки даты подписи.\n"
            "   4.4 Папка табелей - папка, где лежат табели.\n"
            "   4.5 Папка сохранения - куда готовые карточки будут выгружаться.\n"
            "   4.6 Изделие - название изделия, детали которого используем в карточках.\n"
            "   4.7 Месяцы табелей - между какими месяцами происходит случайная выборка.\n"
            "   4.8 Блок ИСКЛЮЧЕНИЯ - если нужен конкретный цех, конкретная профессия, конкретные люди - "
            "(до трёх человек). \nВписывайте, как в табеле (все заглавные), иначе не найдёт! "
            "Это не гарант того, что кого вы впишете - тот будет выбран, так как не в каждом "
            "месяце есть выбранный человек.\n"
            "   4.9 Кнопка ПРОВЕРКА - проверить, соответствие комбинации цех-профессия в маршрутном файле и табеле. "
            "\nТ.е. "
            "если распечатал, например, 304 - 18466, то это значит, что в табеле нет профессии 18466 в цехе 304.\n"
            "Не забудьте выставить год, карточки которого собираетесь делать.\n"
            "\n     Выставите год. Выставите какие детали (по номеру счёта) хотите сделать, хотя при выборе"
            "файла операций - автоматически выставляется максимальное найденое число деталей.\n"
            "   4.10 Нажимаем ПУСК и готово. Через несколько секунд или более (зависит от количества деталей) загориться "
            "розовым справа от кнопки ПРОВЕРКА шкала (означает - готово).\n"
            "   4.11 Нажимаем СТОП и программа закрывается. В папке после выполнения формируется файл Logs.txt - в нём"
            "записываются действия программы. Можно посмотреть, как что и где формируется.\n"
            "\n     В наименование деталей удалите следующие символы: / , - . : ; Это нужно просто для того, чтобы "
            "сохранить файл на вашем ПК.\n"
            "\n     Также если вы хотите изменить или добавить профессию, наблюдателей и альтернативные кодыы, то в "
            "корневой папке лежат три .txt файла. watchers - наблюдатели, proffesions - профессии, "
            "alternative_codes - альтернативные коды."
            "\n         ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ:\n"
            "\nПрограмма работает так. Сначала пользователь выбирает входные данные. Нажимает на кнопки /Активировать/."
            "Когда пользователь выбрал данные – Программа выбирает ячейку в маршрутном файле, где есть слово "
            "/Наименование/ во втором столбце и затем спускается на 3 ячейки ниже, чтобы начать процесс создания"
            "карточки. \n   Таким образом, программа знает все блоки деталей, где есть слово /Наименование/. "
            "Затем начинается цикл: она смотрит на столбец цеха, столбец шифра, столбец оперативного времени "
            "и на наименование операции. \n Шаблон она сразу записывает наименование операции в соответствии с "
            "шифром профессии который написан в маршрутном файле выставляет профессию, однако если в словаре "
            "не будет подходящего шифра профессии она не запишет его. \nПоэтому нужно добавить в словарь нужный шифр "
            "профессии с её наименованием. \nТакже сразу записывается разряд работ. А теперь программа переходит "
            "к самым основным параметрам: это цех и оперативное время. Затем выбирается случайный табель из папки и "
            "относительно цеха и соответственно шифра про профессии и парочку ещё скрытых свойств выбирается "
            "список людей, подходящих под условия. \nДалее программа случайным образом выбирает трёх людей, укоторых "
            "есть пустые ячейки в табеле. \nЗатем она определяет количество пустых ячеек, то есть свободных дней и "
            "берет одну случайную. Таким образом мы имеем трёх случайных людей с тремя случайными датами в те дни, "
            "в которые они были свободны. \nПотом программа в порядке возрастания выставляет даты и людей и записывает"
            " их их в шаблон. После всего этого программа записывает в шаблон даты наблюдения и утверждения с "
            "разницой 2-4 дня.\n"
            "\n     Программа проходит по четырём вариациям поиска."
            "\n1 - совпадение цеха, 11 и 12 кода квалификации, "
            "шифру, свободные дни\n"
            "2 - совпадение цеха, 11 и 12, дополнительный шифр из словаря, свободные дни\n"
            "3 - совпадение цеха, 11и 12, свободные дни\n"
            "4 - совпадение 11 и 12, шифр, свободные дни\n"
            "\nЕсли в первой вариации найдено меньше 3-х человек, то программа берёт их и недостающих из следующего"
            " цикла.\n"
            "\nСпасибо за внимание! Контакты: Щеглов Роман Геннадьевич (ОТиЗ). Номер телефона: 89003095512. E-mail:"
            " hedrafirus@icloud.com"
        )

        label = QLabel(instruction_text)
        font = QFont("Arial", 12)
        label.setFont(font)
        label.setWordWrap(True)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.addWidget(label)
        scroll_area.setWidget(scroll_content)

        layout.addWidget(scroll_area)
        self.setLayout(layout)


class UI(QMainWindow):
    def __init__(self):

        super(UI, self).__init__()
        self.core_Object = Core(self)
        self.Instruction = InstructionDialog
        self.timer = time.time()
        # Load ui file
        uic.loadUi("Chronos_interface.ui", self)

        # dictionary_cypher_and_proffesion(key:value)
        self.alternative_codes = self.load_data("alternative_codes.txt")
        self.proffesions = self.load_data("proffesions.txt")
        logging.info(self.proffesions)
        logging.info(self.alternative_codes)
        self.watchers = self.load_data("watchers.txt")
        # Define widgets

        self.lineEdit_Main = self.findChild(QLineEdit, "lineEdit_Main")
        self.label_5 = self.findChild(QLabel, "label_5")
        self.pushButton_Main = self.findChild(QPushButton, "pushButton_Main")
        self.frame_2 = self.findChild(QFrame, "frame_2")
        self.lineEdit_Marsh = self.findChild(QLineEdit, "lineEdit_Marsh")
        self.pushButton_Marsh = self.findChild(QPushButton, "pushButton_Marsh")
        self.label = self.findChild(QLabel, "label")
        self.frame_4 = self.findChild(QFrame, "frame_4")
        self.lineEdit_Prof = self.findChild(QLineEdit, "lineEdit_Prof")
        self.label_2 = self.findChild(QLabel, "label_2")
        self.pushButton_Prof = self.findChild(QPushButton, "pushButton_Prof")
        self.frame_5 = self.findChild(QFrame, "frame_5")
        self.lineEdit_Rand = self.findChild(QLineEdit, "lineEdit_Rand")
        self.pushButton_Rand = self.findChild(QPushButton, "pushButton_Rand")
        self.frame_6 = self.findChild(QFrame, "frame_6")
        self.lineEdit_Save = self.findChild(QLineEdit, "lineEdit_Save")
        self.pushButton_Save = self.findChild(QPushButton, "pushButton_Save")
        self.label_4 = self.findChild(QLabel, "label_4")
        self.frame = self.findChild(QFrame, "frame")
        self.lineEdit_Watcher2 = self.findChild(QLineEdit, "lineEdit_Watcher2")
        self.label_14 = self.findChild(QLabel, "label_14")
        self.lineEdit_Watcher1 = self.findChild(QLineEdit, "lineEdit_Watcher1")
        self.frame_7 = self.findChild(QFrame, "frame_7")
        self.lineEdit_Inc3 = self.findChild(QLineEdit, "lineEdit_Inc3")
        self.lineEdit_Inc1 = self.findChild(QLineEdit, "lineEdit_Inc1")
        self.label_8 = self.findChild(QLabel, "label_8")
        self.lineEdit_Inc2 = self.findChild(QLineEdit, "lineEdit_Inc2")
        self.label_6 = self.findChild(QLabel, "label_6")
        self.label_7 = self.findChild(QLabel, "label_7")
        self.lineEdit_inc_workshop = self.findChild(QLineEdit, "lineEdit_inc_workshop")
        self.lineEdit_inc_prof = self.findChild(QLineEdit, "lineEdit_inc_prof")
        self.frame_8 = self.findChild(QFrame, "frame_8")
        self.label_15 = self.findChild(QLabel, "label_15")
        self.LogFrame = self.findChild(QTextEdit, "LogFrame")
        self.ProgressBar = self.findChild(QProgressBar, "ProgressBar")
        self.StartButton = self.findChild(QPushButton, "StartButton")
        self.CounterOfStarts = self.findChild(QLCDNumber, "CounterOfStarts")
        self.frame_9 = self.findChild(QFrame, "frame_9")
        self.label_10 = self.findChild(QLabel, "label_10")
        self.label_12 = self.findChild(QLabel, "label_12")
        self.label_11 = self.findChild(QLabel, "label_11")
        self.spinBox_2 = self.findChild(QSpinBox, "spinBox_2")
        self.spinBox = self.findChild(QSpinBox, "spinBox")
        self.frame_10 = self.findChild(QFrame, "frame_10")
        self.label_13 = self.findChild(QLabel, "label_13")
        self.lineEdit_Name = self.findChild(QLineEdit, "lineEdit_Name")
        self.StopButton = self.findChild(QPushButton, "StopButton")
        self.menubar = self.findChild(QMenuBar, "menubar")
        self.statusbar = self.findChild(QStatusBar, "statusbar")
        self.label_16 = self.findChild(QLabel, "label_16")
        self.pushButton_Save_2 = self.findChild(QPushButton, "pushButton_Save_2")
        self.label_9 = self.findChild(QLabel, "label_9")
        self.Check_button = self.findChild(QPushButton, "Check_button")
        self.Spin_Year = self.findChild(QSpinBox, "Spin_Year")
        self.label_year = self.findChild(QLabel, "label_year")
        self.instructionButton = self.findChild(QPushButton, "instructionButton")
        self.spinBox_d1 = self.findChild(QSpinBox, "spinBox_d1")
        self.spinBox_d2 = self.findChild(QSpinBox, "spinBox_d2")
        self.spinBox_space = self.findChild(QSpinBox, "spinBox_space")

        # Define Variables
        self.info_textbox_operations = ""
        self.info_textbox_save = ""
        self.info_textbox_template = ""
        self.info_textbox_tabels = ""
        self.info_textbox_cal = ""
        self.Spin_Year.setValue(2024)

        # Do something
        self.pushButton_Marsh.clicked.connect(self.clickbutton_operations)
        self.pushButton_Save.clicked.connect(self.clickbutton_save)
        self.pushButton_Main.clicked.connect(self.clickbutton_template)
        self.pushButton_Cal.clicked.connect(self.clickbutton_cal)
        self.StartButton.clicked.connect(self.clickbutton_start)
        self.StopButton.clicked.connect(self.clickbutton_stop)
        self.pushButton_Save_2.clicked.connect(self.get_space_tabels)
        self.Check_button.clicked.connect(self.get_check_data)
        self.instructionButton.clicked.connect(self.open_instruction)

        # Set color
        self.label.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_4.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_5.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_16.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_9.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_13.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )
        self.label_year.setStyleSheet(
            "color: rgb(255, 255, 255); background-color: rgba(25, 10, 25, 100);"
            'font: 87 8pt "Arial Black"; border-radius: 7px;'
        )

        # Установка фона
        self.setStyleSheet(
            "#MainWindow{border-image:url(Chronos_background.jpeg);"
            "background-repeat: no-repeat;}"
        )
        self.themes_box.addItems(
            [
                "Chronos_background.jpeg",
                "background_1.WEBP",
                "background_2.WEBP",
                "background_3.jpg",
                "background_4.jpg",
                "background_5.jpg",
            ]
        )
        self.themes_box.currentIndexChanged.connect(self.set_themes)

        # Show app
        self.show()

    def open_instruction(self):
        logging.info("Открыта инструкция")
        dialog = InstructionDialog()
        dialog.initUI()
        dialog.exec_()
        logging.info("Инструкция закрыта")

    #  Load Data_dictionary
    def load_data(self, filepath):
        data = {}
        with open(filepath, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines:
                key, values = line.strip().split(": ")
                key = int(key)
                data[key] = values.split(", ")
        logging.info("Загружен один из словарей")
        return data

    # Setting themes
    def set_themes(self, index):
        selectedtheme = self.themes_box.itemText(index)
        stylesheet = f"#MainWindow {{border-image: url({selectedtheme}); background-repeat: no-repeat;}}"
        self.setStyleSheet(stylesheet)
        logging.info(f"Выбрана тема: {selectedtheme}")

    # Загрузить файл операций и изменить цвет лейбла
    def clickbutton_operations(self):
        self.info_textbox_operations = str(QFileDialog.getOpenFileName(self)[0])
        if self.info_textbox_operations != "":
            try:
                self.wb_operations = opxl.load_workbook(self.info_textbox_operations)
                logging.info(f"Выбран файл операций: {self.info_textbox_operations}")
                self.LogFrame.insertPlainText(
                    f"\nВыбран файл операций: {self.info_textbox_operations}"
                )
                self.label.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    'font: 87 8pt "Arial Black";'
                    "background-color: transparent"
                )
                logging.info(
                    f"\n{self.core_Object.get_way_rows()}<--Количество деталей"
                )
                self.spinBox_d2.setValue(self.core_Object.get_way_rows())
            except:
                logging.info(
                    f"Выбран неверный файл операций {self.info_textbox_operations}"
                )

    # Загрузить путь сохранения и изменить цвет лейбла
    def clickbutton_save(self):
        self.info_textbox_save = str(QFileDialog.getExistingDirectory(self))
        if self.info_textbox_save != "":
            self.LogFrame.insertPlainText(
                f"\nВыбран путь сохранения: {str(self.info_textbox_save)}"
            )
            self.label_4.setStyleSheet(
                "color: rgb(255, 255, 255);"
                'font: 87 8pt "Arial Black";'
                "background-color: transparent"
            )
            logging.info(f"Выбран путь сохранения: {str(self.info_textbox_save)}")

    # Загрузить файл-шаблон и изменить цвет лейбла
    def clickbutton_template(self):
        self.info_textbox_template = str(QFileDialog.getOpenFileName(self)[0])
        if self.info_textbox_template != "":
            try:
                logging.info(f"Выбран файл-шаблон: {self.info_textbox_template}")
                self.LogFrame.insertPlainText(
                    f"\nВыбран шаблон: {str(self.info_textbox_template)}"
                )
                self.label_5.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    'font: 87 8pt "Arial Black";'
                    "background-color: transparent"
                )
            except:
                logging.info("Выбран неверный файл шаблона")

    # Загрузить файл-календарь и изменить цвет лейбла
    def clickbutton_cal(self):
        self.info_textbox_cal = str(QFileDialog.getOpenFileName(self)[0])
        if self.info_textbox_cal != "":
            try:
                self.wb_calendar = opxl.load_workbook(self.info_textbox_cal)
                logging.info(f"Выбран файл календаря: {self.info_textbox_cal}")
                self.LogFrame.insertPlainText(
                    f"\nВыбран календарь: {str(self.info_textbox_cal)}"
                )
                self.label_9.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    'font: 87 8pt "Arial Black";'
                    "background-color: transparent"
                )
            except:
                logging.info("Выбран неверный файл календаря")

    # Загрузить путь к папке табелей и изменить цвет лейбла
    def get_space_tabels(self):
        self.info_textbox_tabels = str(QFileDialog.getExistingDirectory(self))
        if self.info_textbox_tabels != "":
            self.LogFrame.insertPlainText(
                f"\nВыбрана папка табелей: {str(self.info_textbox_tabels)}"
            )
            self.label_16.setStyleSheet(
                "color: rgb(255, 255, 255);"
                'font: 87 8pt "Arial Black";'
                "background-color: transparent"
            )
            logging.info(f"Выбрана папка табелей: {self.info_textbox_tabels}")

    def get_check_data(self):
        if self.core_Object.active_inc() is True:
            self.LogFrame.insertPlainText("\nБлок исключений активен!")
            logging.info("Блок исключений активен")
        else:
            self.LogFrame.insertPlainText("\nБлок исключений НЕактивен!")
            logging.info("Блок исключений НЕактивен")

        if any(
            value == ""
            for value in [self.info_textbox_operations, self.info_textbox_tabels]
        ):
            self.LogFrame.insertPlainText(
                "\nПроверка недоступна! Выберите файл операций и папку табелей."
            )
            logging.info("Операции и табели не выбраны")
        else:
            tab = self.core_Object.get_random_tabel()
            combinations = set()
            combinations_2 = set()
            for row in self.core_Object.wb_tabel[
                self.core_Object.wb_tabel.sheetnames[0]
            ].iter_rows(min_row=2, values_only=True):
                workshop, cipher = row[0], row[8]
                combination = (workshop, cipher)
                combinations.add(combination)
                workshop = row[1]
                try:
                    int(workshop)
                    combination_2 = (workshop, cipher)
                    combinations_2.add(combination_2)
                except:
                    pass
            missing_combinations = set()
            for row in self.wb_operations[self.wb_operations.sheetnames[0]].iter_rows(
                min_row=2, values_only=True
            ):
                workshop, cipher = (
                    row[1],
                    row[3],
                )
                if isinstance(workshop, int):
                    combination = (workshop, cipher)
                    if combination not in combinations and combinations_2:
                        missing_combinations.add(combination)
                if not self.proffesions.get(row[3]) and row[3] is not None:
                    self.LogFrame.insertPlainText(
                        f"\nВ словаре нет профессии ==> {row[3]}"
                    )
                    logging.info(f"\nВ словаре нет профессии ==> {row[3]}")
                if (
                    not self.watchers.get(row[1])
                    and row[1] is not None
                    and type(row[1]) is not str
                ):
                    self.LogFrame.insertPlainText(
                        f"\nВ словаре наблюдателей - нет цеха ==> {row[1]})"
                    )
                    logging.info(f"\nВ словаре наблюдателей - нет цеха ==> {row[1]}")
            if missing_combinations:
                logging.info("Выполнение проверки (сверки)")
                self.LogFrame.insertPlainText(f"\nНет в табеле ({tab}):")
                for combination in missing_combinations:
                    self.LogFrame.insertPlainText(f"\n{combination}")
            else:
                self.LogFrame.insertPlainText(f"\nВсё совпадает!")

    # Сохранить и записать в файл Log, и выйти из приложения
    def clickbutton_stop(self):
        sys.exit()

    # Инициализация функций и так далее по нажатию на кнопку "Пуск"
    def clickbutton_start(self):
        # Переменная для наименования изделия (текст)
        self.name_product = self.lineEdit_Name.text()
        self.year_text = self.Spin_Year.value()
        # Проверки на заполненность полей и изменение цвета виджетов текста
        if any(
            value == ""
            for value in [
                self.info_textbox_operations,
                self.info_textbox_save,
                self.name_product,
                self.info_textbox_template,
                self.info_textbox_tabels,
                self.info_textbox_cal,
            ]
        ):
            self.LogFrame.insertPlainText("\nВыбраны не все элементы!")
            if self.info_textbox_operations == "":
                self.label.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )
            if self.info_textbox_save == "":
                self.label_4.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )
            if self.info_textbox_template == "":
                self.label_5.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )
            if self.name_product == "":
                self.label_13.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )
            if self.info_textbox_tabels == "":
                self.label_16.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )
            if self.info_textbox_cal == "":
                self.label_9.setStyleSheet(
                    "color: rgb(255, 255, 255);"
                    "background-color: rgba(255, 50, 50, 100);"
                    'font: 87 8pt "Arial Black"; border-radius: 7px;'
                )

        else:  # Печатаем в лог информацию о изделии
            self.LogFrame.insertPlainText(
                "\n----------------------------------------------"
                "\nИзделие: " + self.name_product
            )
            logging.info(
                f"\n----------------------------------------------\nИзделие: {self.name_product}"
            )
            self.LogFrame.insertPlainText(
                "\nМесяцы: с "
                + str(self.spinBox.value())
                + " по "
                + str(self.spinBox_2.value())
            )
            logging.info(
                f"Месяцы: с {str(self.spinBox.value())} по {str(self.spinBox_2.value())}"
            )

            if (
                (self.lineEdit_Inc1.text() != "")
                or (self.lineEdit_Inc2.text() != "")
                or (self.lineEdit_Inc3.text() != "")
            ):
                self.LogFrame.insertPlainText(
                    "\nИсключения: "
                    + self.lineEdit_Inc1.text()
                    + ", "
                    + self.lineEdit_Inc2.text()
                    + ", "
                    + self.lineEdit_Inc3.text()
                    + ", "
                    + "\nЦех (исключение): "
                    + self.lineEdit_inc_workshop.text()
                    + "\nКод профессии (исключение): "
                    + self.lineEdit_inc_prof.text()
                )
                logging.info(
                    f"Исключения: {self.lineEdit_Inc1.text()}, {self.lineEdit_Inc2.text()}, "
                    f"{self.lineEdit_Inc3.text()}\n по цеху {self.lineEdit_inc_workshop.text()} с "
                    f"кодом {self.lineEdit_inc_prof.text()}"
                )
            else:
                self.LogFrame.insertPlainText("\nНет исключений")
                logging.info("Без исключений")

            # Adds

            logging.info(
                f"{self.core_Object.get_count_of_space()}----Всего строк в файле операций"
            )
            self.ProgressBar.setValue(25)
            logging.info(f"\n{self.core_Object.get_way_rows()}----Количество деталей")
            self.LogFrame.insertPlainText(
                f"\nСписок стартовых строк => {self.core_Object.list_startrows_of_things}"
            )
            logging.info(
                f"Список стартовых строк => {self.core_Object.list_startrows_of_things}"
            )

            self.core_Object.make_buty()
            self.ProgressBar.setValue(100)


class Core:
    def __init__(self, interface_Object):
        self.interface_Object = interface_Object

    # Генератор случайных чисел с отклонением в 20 процентов
    def generate_values(self):
        self.values = [
            round(random.uniform(0.9 * self.target_mean, 1.1 * self.target_mean), 5)
            for _ in range(3)
        ]
        logging.info(self.values)
        # Текущее среднее
        current_mean = sum(self.values) / 3
        # Масштабирум значения, чтобы среднее значение было равно целевому
        scaled_values = [
            round(value * self.target_mean / current_mean, 5) for value in self.values
        ]
        return scaled_values

    # Вспомогательная функция к initialize_exel, которая генерирует строку чисел по правилам и перемешивает её
    def initialize_exel_2(self):
        current = random.uniform(1.03, 1.09)
        percents = [1, current]

        for _ in range(8 - 2):
            min_border = max(1 / current, 0.9)
            max_border = min(percents[1] / current, 1.1)

            current *= random.uniform(min_border, max_border)
            percents.append(current)

        coeff = self.middle / sum(percents) * len(percents)
        result = [round(percent * coeff, 5) for percent in percents]
        self.coef_chrono = round(max(result) / min(result), 2)
        y = random.randint(972, 990) / 1000
        minimal_ghost_value = round(min(result) * y, 5)
        logging.info(minimal_ghost_value, "<-Минимальное гостовое")
        a = (random.randint(11150, 15988) / 100000) + 1
        x = min(result) * y * a / max(result)
        maximal_ghost_value = round(max(result) * x, 5)
        logging.info(maximal_ghost_value, "<-Максимальное гостовое")
        logging.info("Высокий множитель:", x, "Низкий множитель:", y)
        logging.info(a, "<-Заданный предел")
        logging.info(round(maximal_ghost_value / minimal_ghost_value, 2))
        result.append(minimal_ghost_value)
        result.append(maximal_ghost_value)
        random.shuffle(result)
        return result

    # Заполняет числовые поля относительно target_mean(Топер) в шаблон
    def initialize_exel(self):
        temporary_massiv = []
        temporary_massiv = self.generate_values()
        counter = 0
        for i in range(15, 22, 3):
            self.wb_template[self.wb_template.sheetnames[-1]].cell(
                row=i, column=14, value=round(temporary_massiv[counter], 5)
            )
            temporary_massiv_2 = []
            # self.target_mean = float(temporary_massiv[counter])
            self.middle = temporary_massiv[counter]
            temporary_massiv_2 = self.initialize_exel_2()
            self.wb_template[self.wb_template.sheetnames[-1]].cell(
                row=i, column=15, value=self.coef_chrono
            )
            for j in range(1, 11):
                self.wb_template[self.wb_template.sheetnames[-1]].cell(
                    row=i, column=j + 3, value=temporary_massiv_2[j - 1]
                )
            counter += 1

    # Возвращает общее количество строк в файле операций
    def get_count_of_space(self):
        total = self.interface_Object.wb_operations[
            self.interface_Object.wb_operations.sheetnames[0]
        ].max_row
        return total

    # Возвращает количество деталей (считает строки начинающиеся с "Наименование") в файле операций
    def get_way_rows(self):
        self.NamesOfThings = []
        self.list_startrows_of_things = []
        s = 0
        j = self.get_count_of_space()
        for i in range(1, j + 1):
            if str(
                self.interface_Object.wb_operations[
                    self.interface_Object.wb_operations.sheetnames[0]
                ]
                .cell(row=i, column=2)
                .value
            ).startswith("Наименование узла (детали)"):
                s += 1
                self.NamesOfThings.append(i)
                self.list_startrows_of_things.append(
                    i + self.interface_Object.spinBox_space.value()
                )  # i+3 это разница между наименованием и началом строк данных
        return s

    # Загружает в память и возвращает случайный табель из папки в интервале SpinBox-SpinBox_2
    def get_random_tabel(self):
        self.month = random.randint(
            self.interface_Object.spinBox.value(),
            self.interface_Object.spinBox_2.value(),
        )
        tabel = f"Tab{self.month}.xlsx"
        self.wb_tabel = opxl.load_workbook(
            f"{self.interface_Object.info_textbox_tabels}/{tabel}"
        )
        return tabel

    # Возвращает номера колонок, которые соответствуют пустым ячейкам; Prime_rows - массив взятых людей
    def get_random_day(self, row_number):
        Temp_rows = []
        counter = 1
        for i in range(31, 93, 2):
            if (
                self.wb_tabel[self.wb_tabel.sheetnames[0]]
                .cell(row=row_number + 2, column=i)
                .value
                is not None
            ) and (
                self.wb_tabel[self.wb_tabel.sheetnames[0]]
                .cell(row=row_number + 2, column=i + 1)
                .value
                is None
            ):

                Temp_rows.append(counter)
            counter += 1
        return Temp_rows

    def active_inc(self):
        if (
            (self.interface_Object.lineEdit_inc_prof.text() != "")
            and (self.interface_Object.lineEdit_inc_workshop != "")
            and (
                any(
                    [
                        self.interface_Object.lineEdit_Inc1.text(),
                        self.interface_Object.lineEdit_Inc2.text(),
                        self.interface_Object.lineEdit_Inc3.text(),
                    ]
                )
            )
        ):
            return True
        else:
            return False

    def get_people_original(self):
        people = []
        people_2 = []
        people_3 = []
        people_4 = []
        people_inc = []
        for row_number, row in enumerate(
            self.wb_tabel[self.wb_tabel.sheetnames[0]].iter_rows(min_row=2)
        ):
            flag = True
            try:
                if (
                    (
                        row[0].value == self.workshop
                        or int(row[1].value) == int(self.workshop)
                    )
                    and (row[7].value == 11 or row[7].value == 12)
                    and row[8].value == self.cypher
                    and row[0].value != 250
                    and self.get_random_day(row_number)
                ):

                    people.append(row_number + 2)
                    flag = False

            except:
                pass

            if self.interface_Object.alternative_codes.get(self.cypher):
                for addcypher in self.interface_Object.alternative_codes.get(
                    self.cypher
                ):
                    try:
                        if (
                            (
                                row[0].value == self.workshop
                                or int(row[1].value == int(self.workshop))
                            )
                            and (row[7].value == 11 or row[7].value == 12)
                            and row[8].value == int(addcypher)
                            and row[0].value != 250
                            and self.get_random_day(row_number)
                        ):
                            people_2.append(row_number + 2)
                            flag = False
                    except:
                        pass

            try:
                if (
                    (
                        row[0].value == self.workshop
                        or int(row[1].value) == int(self.workshop)
                    )
                    and (row[7].value == 11 or row[7].value == 12)
                    and row[0].value != 250
                    and self.get_random_day(row_number)
                    and flag == True
                ):

                    people_3.append(row_number + 2)
            except:
                pass

            try:
                if (
                    (row[7].value == 11 or row[7].value == 12)
                    and row[0].value != 250
                    and self.get_random_day(row_number)
                    and row[8].value == self.cypher
                ):

                    people_4.append(row_number + 2)
            except:
                pass

            if (
                (row[7].value == 11 or row[7].value == 12)
                and self.active_inc()
                and self.get_random_day(row_number)
                and self.interface_Object.lineEdit_inc_workshop.text()
                == str(self.workshop)
                and self.interface_Object.lineEdit_inc_prof.text() == str(self.cypher)
                and row[8].value == self.cypher
                and row[5].value
                in [
                    self.interface_Object.lineEdit_Inc1.text(),
                    self.interface_Object.lineEdit_Inc2.text(),
                    self.interface_Object.lineEdit_Inc3.text(),
                ]
            ):
                people_inc.append(row_number + 2)
        return [people, people_2, people_3, people_4, people_inc]

    #  Цикл поиска людей по шифру в Острогоржске
    def get_people_Ostrog(self):
        people = []
        people_2 = []
        people_3 = []
        people_4 = []
        people_inc = []
        for row_number, row in enumerate(
            self.wb_tabel[self.wb_tabel.sheetnames[0]].iter_rows(min_row=2)
        ):
            flag = True
            if (
                row[0].value == self.workshop
                and (row[7].value == 11 or row[7].value == 12)
                and row[8].value == self.cypher
                and self.get_random_day(row_number)
            ):
                people.append(row_number + 2)
                flag = False

            if self.interface_Object.alternative_codes.get(self.cypher):
                for addcypher in self.interface_Object.alternative_codes.get(
                    self.cypher
                ):
                    try:
                        if (
                            row[0].value == self.workshop
                            and (row[7].value == 11 or row[7].value == 12)
                            and row[8].value == addcypher
                            and row[0].value == 250
                            and self.get_random_day(row_number)
                        ):
                            people_2.append(row_number + 2)
                            flag = False
                    except:
                        pass

            try:
                if (
                    row[0].value == self.workshop
                    and (row[7].value == 11 or row[7].value == 12)
                    and self.get_random_day(row_number)
                    and flag == True
                ):

                    people_3.append(row_number + 2)
            except:
                pass

            try:
                if (
                    (row[7].value == 11 or row[7].value == 12)
                    and self.get_random_day(row_number)
                    and row[8].value == self.cypher
                ):

                    people_4.append(row_number + 2)
            except:
                pass

            if (
                (row[7].value == 11 or row[7].value == 12)
                and self.active_inc()
                and self.get_random_day(row_number)
                and self.workshop == 250
                and self.interface_Object.lineEdit_inc_workshop.text()
                == str(self.workshop)
                and self.interface_Object.lineEdit_inc_prof.text() == str(self.cypher)
                and row[8].value == self.cypher
                and row[5].value
                in [
                    self.interface_Object.lineEdit_Inc1.text(),
                    self.interface_Object.lineEdit_Inc2.text(),
                    self.interface_Object.lineEdit_Inc3.text(),
                ]
            ):
                people_inc.append(row_number + 2)
        return [people, people_2, people_3, people_4, people_inc]

    def make_new_list(self):
        new_sheet = self.wb_template.copy_worksheet(
            self.wb_template[self.wb_template.sheetnames[-1]]
        )
        # Устанавливаем новое имя для нового листа
        new_sheet.title = f"{self.operation}"

    def get_watcher(self):
        try:
            watcher = random.choice(self.interface_Object.watchers.get(self.workshop))
        except:
            watcher = "-ПУСТО-"
        return watcher

    def get_calendar_day(self, current_date):
        logging.info(current_date)
        for cycle in range(1, 3):
            rand_val = random.randint(cycle, 4)
            current_date += timedelta(days=rand_val)
            logging.info(current_date)
            month_row = current_date.month + 12 * (current_date.year - 2021)
            while (
                self.interface_Object.wb_calendar[
                    self.interface_Object.wb_calendar.sheetnames[-1]
                ]
                .cell(row=month_row, column=current_date.day)
                .value
                is not None
            ):
                current_date += timedelta(days=1)
                month_row = current_date.month + 12 * (current_date.year - 2021)
            self.wb_template[self.wb_template.sheetnames[-1]].cell(
                row=28 + cycle,
                column=4,
                value=current_date.strftime("%d.%m.%Y"),
            )

    # Функция возвращает наименования относительно спика строк NamesOfThings (начинающиеся с "Наименование") и делает
    # дополнительные иттеративные действия
    def make_buty(self):
        # Добавить в цикл на каждый лист!!! len(self.NamesOfThings)
        # Новое изделие на новом листе
        for makes in range(
            self.interface_Object.spinBox_d1.value() - 1,
            self.interface_Object.spinBox_d2.value(),
        ):
            Name_thing = str(
                self.interface_Object.wb_operations[
                    self.interface_Object.wb_operations.sheetnames[0]
                ]
                .cell(row=self.NamesOfThings[makes], column=2)
                .value
            )[46:]
            logging.info(f"\nНаименования =>{Name_thing}")
            tabel_text = self.get_random_tabel()
            self.wb_template = opxl.load_workbook(
                self.interface_Object.info_textbox_template
            )
            logging.info(f"\n{tabel_text}<-Выбранный табель")
            startcell = self.list_startrows_of_things[makes]
            # Цикл для заполнения и создания листов деталей в изделии
            while (
                self.interface_Object.wb_operations[
                    self.interface_Object.wb_operations.sheetnames[0]
                ]
                .cell(row=startcell, column=6)
                .value
                is not None
            ):
                if (
                    self.interface_Object.wb_operations[
                        self.interface_Object.wb_operations.sheetnames[0]
                    ]
                    .cell(row=startcell, column=6)
                    .value
                    != 0
                    and self.interface_Object.wb_operations[
                        self.interface_Object.wb_operations.sheetnames[0]
                    ]
                    .cell(row=startcell, column=2)
                    .value
                    is not None
                ):
                    # Присваиваю значения со строки (цех, операцию, шифр, разряд, Топер)
                    self.workshop = int(
                        self.interface_Object.wb_operations[
                            self.interface_Object.wb_operations.sheetnames[0]
                        ]
                        .cell(row=startcell, column=2)
                        .value
                    )
                    self.operation = (
                        self.interface_Object.wb_operations[
                            self.interface_Object.wb_operations.sheetnames[0]
                        ]
                        .cell(row=startcell, column=3)
                        .value
                    )
                    self.cypher = (
                        self.interface_Object.wb_operations[
                            self.interface_Object.wb_operations.sheetnames[0]
                        ]
                        .cell(row=startcell, column=4)
                        .value
                    )
                    self.work_number = (
                        self.interface_Object.wb_operations[
                            self.interface_Object.wb_operations.sheetnames[0]
                        ]
                        .cell(row=startcell, column=5)
                        .value
                    )
                    self.timing = (
                        self.interface_Object.wb_operations[
                            self.interface_Object.wb_operations.sheetnames[0]
                        ]
                        .cell(row=startcell, column=6)
                        .value
                    )

                    self.make_new_list()  # Создаю копию последнего листа
                    # Беру оперативное время
                    self.target_mean = float(
                        (
                            self.interface_Object.wb_operations[
                                self.interface_Object.wb_operations.sheetnames[0]
                            ]
                            .cell(row=startcell, column=6)
                            .value
                        )
                    )
                    self.initialize_exel()  # Заполняем таблицу рандомными числами

                    #  Логирую и тд
                    try:
                        if self.workshop == 250:
                            special_list = []
                            ways_ost = self.get_people_Ostrog()
                            logging.info(f"\nВыбранные люди Острог => {ways_ost}")
                            if self.active_inc():
                                k = -1
                            else:
                                k = 0
                            used_days = []
                            while len(special_list) < 3:
                                if ways_ost[k]:
                                    for sublist in ways_ost[k]:
                                        temp_way = random.choice(ways_ost[k])
                                        days = self.get_random_day(
                                            row_number=temp_way - 2
                                        )
                                        if used_days:
                                            for subcount in used_days:
                                                try:
                                                    days.remove(subcount)
                                                except:
                                                    pass
                                        logging.info(
                                            days,
                                            "Дни(days)",
                                            temp_way,
                                            "<-Строка человека",
                                        )
                                        used_day = random.choice(days)
                                        used_days.append(used_day)
                                        special_list.append(
                                            [
                                                temp_way,
                                                used_day,
                                            ]
                                        )
                                        ways_ost[k].remove(temp_way)
                                        if len(special_list) == 3:
                                            break
                                    if k >= 0 and self.active_inc():
                                        self.interface_Object.LogFrame.insertPlainText(
                                            f"\nПрошли не все исключения в текущем месяце!"
                                        )
                                        logging.info(
                                            "Прошли не все исключения в текущем месяце"
                                        )
                                k += 1
                            special_list.sort(key=lambda x: x[1])
                            logging.info(
                                special_list, "<==Отсортированный по возрастанию"
                            )
                            # Логируем специальный список
                            logging.info(f"\nSpecial list: {special_list}\n")
                        else:
                            ways = self.get_people_original()
                            logging.info(f"\nВыбранные люди => {ways}")
                            if self.active_inc():
                                k = -1
                            else:
                                k = 0
                            special_list = []
                            while len(special_list) < 3:
                                if ways[k]:
                                    for sublist in ways[k]:
                                        temp_way = random.choice(ways[k])
                                        days = self.get_random_day(
                                            row_number=temp_way - 2
                                        )
                                        logging.info(days, "Дни(days)")
                                        special_list.append(
                                            [
                                                temp_way,
                                                random.choice(days),
                                            ]
                                        )
                                        ways[k].remove(temp_way)
                                        if len(special_list) == 3:
                                            break
                                    if k >= 0 and self.active_inc():
                                        self.interface_Object.LogFrame.insertPlainText(
                                            f"\nПрошли не все исключения в текущем месяце!"
                                        )
                                        logging.info(
                                            "Прошли не все исключения в текущем месяце"
                                        )
                                k += 1

                            special_list.sort(key=lambda x: x[1])
                            logging.info(special_list, "<===Отсортированный")
                            # Логируем специальный список
                            logging.info(f"\nSpecial list: {special_list}")
                    except:
                        logging.info(f"Пустой массив людей? Строка: {startcell}")
                        self.interface_Object.LogFrame.insertPlainText(
                            f"\nНе нашёлся цех по строчке {startcell}"
                            f"\nЯчейка: {startcell}, Цех: {self.workshop}, Операция: {self.operation}, Шифр: {self.cypher}, Разряд: {self.work_number}, Топер: {self.timing}"
                        )
                        special_list = []
                    if special_list:
                        # Записываю в шаблон данные
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "D7"
                        ] = self.interface_Object.lineEdit_Name.text()  # Изделие
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "F8"
                        ] = Name_thing  # Деталь
                        logging.info(Name_thing, "---", self.operation)
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "D9"
                        ] = self.operation  # Операция
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "D10"
                        ] = self.work_number  # Разряд
                        # Профессия
                        logging.info(self.interface_Object.proffesions.get(self.cypher))
                        if self.interface_Object.proffesions.get(self.cypher):
                            self.wb_template[self.wb_template.sheetnames[-1]][
                                "B12"
                            ].value = self.interface_Object.proffesions.get(
                                self.cypher
                            )[
                                0
                            ]
                        else:
                            self.interface_Object.LogFrame.insertPlainText(
                                f"\n{self.cypher}->Нет профессии в словаре"
                            )
                            logging.info(f"\n{self.cypher}->Нет профессии в словаре")
                            self.wb_template[self.wb_template.sheetnames[-1]][
                                "B12"
                            ] = self.cypher
                        #  Наблюдатели
                        watcher_temp = self.get_watcher()
                        logging.info(
                            f"\n{watcher_temp}<-Наблюдатель по {self.workshop} цеху"
                        )
                        if watcher_temp == "No watcher":
                            self.interface_Object.LogFrame.insertPlainText(
                                f"\n===На листе {self.wb_template.sheetnames[-1]} в детали {Name_thing} нет наблюдателя!"
                            )
                            logging.info(
                                f"\n===На листе {self.wb_template.sheetnames[-1]} в детали {Name_thing} нет наблюдателя!"
                            )
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "D6"
                        ] = watcher_temp
                        self.wb_template[self.wb_template.sheetnames[-1]][
                            "J29"
                        ] = watcher_temp
                        counter = 0
                        for worker_number in range(15, 22, 3):
                            if special_list:
                                worker = (
                                    self.wb_tabel[self.wb_tabel.sheetnames[0]]
                                    .cell(row=special_list[counter][0], column=6)
                                    .value
                                )
                            else:
                                worker = "--ПУСТО--"
                            self.wb_template[self.wb_template.sheetnames[-1]].cell(
                                row=worker_number, column=1, value=worker
                            )
                            counter += 1
                            #  Дни наблюдения
                        for day_range in range(3, 6):
                            if special_list:
                                current_date = datetime.strptime(
                                    f"{special_list[day_range - 3][1]}.{self.month}.{self.interface_Object.year_text}",
                                    "%d.%m.%Y",
                                )
                                logging.info(special_list, "<== SPECIAL LIST в конце")
                                logging.info(
                                    special_list[day_range - 3][1],
                                    "<--Special столбец дня",
                                )
                                self.wb_template[self.wb_template.sheetnames[-1]].cell(
                                    row=day_range,
                                    column=8,
                                    value=current_date.strftime("%d.%m.%Y"),
                                )
                        self.get_calendar_day(current_date)
                    else:
                        del self.wb_template[self.wb_template.sheetnames[-1]]

                startcell += 1
            del self.wb_template[self.wb_template.sheetnames[0]]
            self.wb_template.save(
                f"{self.interface_Object.info_textbox_save}/{Name_thing}.xlsx"
            )


app = QApplication(sys.argv)  # Присваивание переменной действий приложений
UIWindow = UI()
app.exec_()
